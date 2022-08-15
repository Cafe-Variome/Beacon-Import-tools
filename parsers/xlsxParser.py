import json
import openpyxl


class responseParser:

	def __init__(self, doc_file):
		"""
		takes in a xlsx spreadsheet to parse, values can be returned as a JSON or dict structure
		:param doc_file: xlsx spreadsheet
		"""
		print(doc_file)
		wb = openpyxl.load_workbook(doc_file)
		sheet = wb.active
		self.id = self.get_ejp_id(sheet)
		self.name = self.get_name(sheet)
		self.description = self.get_description(sheet)
		self.homepage = self.get_homepage(sheet)
		self.resourceTypes = self.get_resourceTypes(sheet)
		self.apiVersions = self.get_apiVersions(sheet)

	def get_ejp_id(self, sheet):
		"""
		Returns the EJP ID
		:param sheet: xlsx workbook sheet
		:return: id
		"""
		return sheet.cell(row=2, column=2).value

	def get_name(self, sheet):
		"""
		returns the resources name
		:param sheet: xlsx workbook sheet
		:return: string
		"""
		return sheet.cell(row=3, column=2).value

	def get_description(self, sheet):
		"""
		returns description
		:param sheet: xlsx workbook sheet
		:return: string
		"""
		return sheet.cell(row=4, column=2).value

	def get_homepage(self, sheet):
		"""
		returns main URL of resource
		:param sheet: xlsx workbook sheet
		:return: string
		"""
		return sheet.cell(row=5, column=2).value

	def get_resourceTypes(self, sheet):
		"""
		returns a list of resourceTypes
		:param sheet: xlsx workbook sheet
		:return: list
		"""
		resourceTypes = []
		if sheet.cell(row=10, column=2).value is not None:
			resourceTypes.append("PatientRegistryDataset")
		if sheet.cell(row=11, column=2).value is not None:
			resourceTypes.append("BiobankDataset")
		if sheet.cell(row=12, column=2).value is not None:
			resourceTypes.append("KnowledgeBase")
		if sheet.cell(row=13, column=2).value is not None:
			resourceTypes.append("Catalogue")
		return resourceTypes

	def get_apiVersions(self, sheet):
		"""
		returns all apiVersions as a dict
		:param sheet: xlsx workbook sheet
		:return: dict
		"""
		apiVersions = {"queryBuilder": {}, "Beacon": {}}
		if sheet.cell(row=6, column=2).value is not None:
			apiVersions["queryBuilder"]["v1"] = sheet.cell(row=6, column=2).value
		if sheet.cell(row=7, column=2).value is not None:
			apiVersions["queryBuilder"]["v2"] = sheet.cell(row=7, column=2).value
		if sheet.cell(row=8, column=2).value is not None:
			apiVersions["queryBuilder"]["v3"] = sheet.cell(row=8, column=2).value
		if sheet.cell(row=9, column=2).value is not None:
			apiVersions["Beacon"]["v2.0.0"] = sheet.cell(row=9, column=2).value
		return apiVersions

	def to_json(self):
		"""
		return this object as a JSON structure
		:return: JSON
		"""
		return json.dumps(self.to_dict())

	def to_dict(self):
		"""
		return this object as a python dict structure
		OPTIONAL: delete keys before returning (none added here)
		:return: dict
		"""
		base = dict(self.__dict__)
		return base
